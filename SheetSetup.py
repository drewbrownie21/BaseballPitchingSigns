#Imports
import os
import openpyxl
from openpyxl.utils import get_column_letter

#For fill color
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.styles import Alignment #For text alignment

FASTBALL_FILL_COLOR = '79EA1C'
CHANGEUP_FILL_COLOR = 'ECE024'
CURVEBALL_FILL_COLOR = 'FA5252'

class GeneralSheet:
    '''
    General Sheet Commands - Open, Clear, Save
    '''

    def check_workbook_exists(self, file_path):
        return os.path.isfile(file_path)

    def open_workbook(self, file_name):
        if GeneralSheet().check_workbook_exists(file_name):
            wb = openpyxl.load_workbook(file_name)
        else:
            wb = openpyxl.Workbook()
            wb.active.title = "Player"
            wb.create_sheet("Coach")
        GeneralSheet().clear_sheet(wb['Player'])
        GeneralSheet().clear_sheet(wb['Coach'])
        return wb

    def clear_sheet(self, sheet_name):
        sheet_name.delete_cols(1, 200)
        sheet_name.delete_cols(1, 200)

    def save_workbook(self, wb, path):
        wb.save(path)

    def generateAllWorkbookFormatting(self, wb, file_path):
        PlayerSheet().main_player_setup(wb)
        CoachSheet(file_path).print_coach_headers(wb)

class PlayerSheet:
    '''
    Create the Player Sheet template
    '''
    # Function to set border
    def set_border(self, ws, cell_range):
        rows = ws[cell_range]
        side = Side(border_style='thin', color="FF000000")

        rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
        max_y = len(rows) - 1  # index of the last row
        for pos_y, cells in enumerate(rows):
            max_x = len(cells) - 1  # index of the last cell
            for pos_x, cell in enumerate(cells):
                border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom
                )
                if pos_x == 0:
                    border.left = side
                if pos_x == max_x:
                    border.right = side
                if pos_y == 0:
                    border.top = side
                if pos_y == max_y:
                    border.bottom = side

                # set new border only if it's one of the edge cells
                if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                    cell.border = border

    def table_col_row_values(self, wb):
        playersheet = wb['Player']
        # Set column width
        column = 1
        while column < 80:
            i = get_column_letter(column)
            playersheet.column_dimensions[i].width = 4
            column += 1
        top_row = ["00", "01", "02", "03", "04", "05", "10", "11", "12", "13", "14", "15", "20", "21", "22", "23", "24",
                  "25"]
        mid_row = ["30", "31", "32", "33", "34", "35", "40", "41", "42", "43", "44", "45", "50", "51", "52", "53", "54",
                  "55"]
        first_col = ["0", "1", "2", "3", "4", "5", "", "0", "1", "2", "3", "4", "5"]
        playersheet.insert_rows(8)

        # Print top row and mid row
        for i in range(len(top_row)):
            playersheet.cell(1, i + 2).value = top_row[i]
            playersheet.cell(8, i + 2).value = mid_row[i]

        # Print first column
        for i in range(len(first_col)):
            playersheet.cell(i + 2, 1).value = first_col[i]

    def add_player_table_color(self, wb):
        playersheet = wb['Player']
        
        fastball_fill_cells = PatternFill(start_color=FASTBALL_FILL_COLOR, end_color=FASTBALL_FILL_COLOR, fill_type='solid')
        changeup_fill_cells = PatternFill(start_color=CHANGEUP_FILL_COLOR, end_color=CHANGEUP_FILL_COLOR, fill_type='solid')
        curveball_fill_cells = PatternFill(start_color=CURVEBALL_FILL_COLOR, end_color=CURVEBALL_FILL_COLOR, fill_type='solid')

        fastball_key_values = ['FO', 'FI', 'FU', 'BF']
        changeup_key_values = ['CI', 'CO']
        curveball_key_values = ['RI', 'RO']

        for i in range(1, 20):
            for j in range(1, 15):
                if playersheet[str(get_column_letter(i)) + str(j)].value in fastball_key_values:
                    playersheet[str(get_column_letter(i)) + str(j)].fill = fastball_fill_cells
                # Changeup Color
                elif playersheet[str(get_column_letter(i)) + str(j)].value in changeup_key_values:
                    playersheet[str(get_column_letter(i)) + str(j)].fill = changeup_fill_cells
                # Curve color
                elif playersheet[str(get_column_letter(i)) + str(j)].value in curveball_key_values:
                    playersheet[str(get_column_letter(i)) + str(j)].fill = curveball_fill_cells


    def main_player_setup(self, wb):
        player_sheet = wb['Player']
        PlayerSheet().table_col_row_values(wb)
        PlayerSheet().set_border(player_sheet, 'A1:C5')
        for i in range(1, 20):
            for j in range(1, 15):
                PlayerSheet().set_border(player_sheet, str(get_column_letter(i)) + str(j) + ':' + str(get_column_letter(i)) + str(j))

        for row in range(1, 25):
            for col in range(1, 20):
                player_sheet.cell(row, col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

class CoachSheet:
    '''
    Create the template for the coaching sheet
    '''
    def __init__(self, file_path):
        self.file_path = file_path

    def print_coach_headers(self, wb):
        coach_sheet = wb['Coach']
        pitches = ["FI", "FO", "FU", "CI", "CO", "RI", "RO", "BF", "PO"]
        picks = ["P1", "P2", "P3", "P4", "P5", "P6", "P7", "P8"]
        plays = ["W", "B1", "B3", "H", "EA", "EP", "HD", "31"]

        # Print headers
        for i in range(len(pitches)):
            coach_sheet.cell(1, i + 1).value = pitches[i]
        for i in range(len(picks)):
            coach_sheet.cell(23, i + 4).value = picks[i]
        for i in range(len(plays)):
            coach_sheet.cell(28, i + 4).value = plays[i]
